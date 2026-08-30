from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from random import randint, sample, choice
from json import load
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]
ROOT = Path(__file__).resolve().parent

ACHIEVEMENTS_FILE = (PROJECT_ROOT / "app" / "src" / "main" / "res" / "raw" / "achievements.json")

OUTPUT_FOLDER = ROOT / "dummy_data"
OUTPUT_FOLDER.mkdir(exist_ok=True)
OUTPUT_FILE = OUTPUT_FOLDER / "dummy_data.xlsx"

DAYS = 360
AVERAGE_CIGARETTES_PER_DAY = (8, 10)
PROBABILITY = 2
SECONDS_IN_DAY = 60 * 60 * 24
DATE_FORMAT = "%Y-%m-%d %H:%M:%S"


@dataclass
class Statistics:
    history: int = 0
    achievements: int = 0
    costs: int = 0
    notes: int = 0

    def summary(self) -> None:
        print("\n".join((
            "",
            "=" * 60,
            "SUMMARY",
            "=" * 60,
            f"History records:    {self.history}",
            f"Achievements:       {self.achievements}",
            f"Cost records:       {self.costs}",
            f"Notes:              {self.notes}",
            "",
            f"Output file:        {OUTPUT_FILE}",
            "=" * 60,
            "Dummy data generation completed successfully.",
            "=" * 60
        )))


def generate_history_day(date: datetime) -> list[dict]:
    if randint(1, 100) <= PROBABILITY:
        return []

    now = datetime.now().replace(microsecond=0)
    start_time = date.replace(hour=7, minute=0, second=0)
    end_time = now if date.date() == now.date() else date.replace(hour=22, minute=59, second=59)

    history = [{
        "Lent": 1 if randint(1, 100) < PROBABILITY else 0,
        "CreatedAt": (start_time + timedelta(seconds=randint(0, int((end_time - start_time).total_seconds())))).strftime(DATE_FORMAT)
    } for _ in range(randint(*AVERAGE_CIGARETTES_PER_DAY))]
    return sorted(history, key=lambda entry: entry["CreatedAt"])


def get_duration(value: int, unit: str) -> timedelta:
    return {
        "DAYS": timedelta(days=value),
        "WEEKS": timedelta(weeks=value),
        "MONTHS": timedelta(days=value * 30),
        "YEARS": timedelta(days=value * 365)
    }[unit]


def update_last_achieved(achievement: dict, achieved_at: datetime) -> None:
    last_achieved = achievement["LastAchieved"]
    if last_achieved is None or achieved_at > datetime.strptime(last_achieved, DATE_FORMAT):
        achievement["LastAchieved"] = achieved_at.strftime(DATE_FORMAT)


def generate_history() -> list[dict]:
    history = []
    start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=DAYS - 1)

    for day in range(DAYS):
        history.extend(generate_history_day(start_date + timedelta(days=day)))
    return history


def generate_achievements(history: list[dict]) -> list[dict]:
    with ACHIEVEMENTS_FILE.open("r", encoding="utf-8") as file:
        data = load(file)

    achievements = [
        {
            "Value": item["value"],
            "Times": 0,
            "LastAchieved": None,
            "Reset": True,
            "Notify": True,
            "Category": category,
            "Unit": item.get("unit", "CIGARETTES"),
            "Id": index,
        }
        for index, (category, item) in enumerate(((category, item) for category, items in data.items() for item in items), start=1)
    ]

    timestamps = sorted(datetime.strptime(entry["CreatedAt"], DATE_FORMAT) for entry in history)

    if not timestamps:
        return achievements

    now = datetime.now().replace(microsecond=0)
    average_cigarettes_per_day = len(timestamps) / max((timestamps[-1] - timestamps[0]).total_seconds() / SECONDS_IN_DAY, 1)

    periods = list(zip(timestamps, timestamps[1:]))
    periods.append((timestamps[-1], now))

    for achievement in achievements:
        threshold = achievement["Value"]

        match achievement["Category"]:
            case "SMOKE_FREE_TIME":
                duration = get_duration(threshold, achievement["Unit"])
                for start, end in periods:
                    if end - start >= duration:
                        achievement["Times"] += 1
                        update_last_achieved(achievement, start + duration)
                achievement["Reset"] = (now - timestamps[-1] < duration)

            case "CIGARETTES_AVOIDED":
                for start, end in periods:
                    smoke_free_days = (end - start).total_seconds() / SECONDS_IN_DAY
                    cigarettes_avoided = (smoke_free_days * average_cigarettes_per_day)

                    if cigarettes_avoided >= threshold:
                        achievement["Times"] += 1
                        update_last_achieved(achievement, start + timedelta(days=threshold / average_cigarettes_per_day))

                current_cigarettes_avoided = ((now - timestamps[-1]).total_seconds() / SECONDS_IN_DAY * average_cigarettes_per_day)
                achievement["Reset"] = (current_cigarettes_avoided < threshold)
    return achievements


def generate_costs() -> list[dict]:
    now = datetime.now().replace(microsecond=0)
    start_date = (now.replace(hour=0, minute=0, second=0) - timedelta(days=DAYS - 1))
    end_date = now.replace(hour=23, minute=59, second=59)

    prices = [0.18, 0.20, 0.21, 0.22, 0.23, 0.24, 0.26, 0.28, 0.30]

    costs = []
    current_start = start_date

    while current_start < end_date:
        remaining_seconds = int((end_date - current_start).total_seconds())

        if remaining_seconds <= 30 * SECONDS_IN_DAY:
            current_end = end_date
        else:
            interval_seconds = randint(30 * SECONDS_IN_DAY, min(120 * SECONDS_IN_DAY, remaining_seconds))
            current_end = current_start + timedelta(seconds=interval_seconds)

        costs.append({
            "Price": choice(prices),
            "StartDate": current_start.strftime(DATE_FORMAT),
            "EndDate": current_end.strftime(DATE_FORMAT),
        })
        current_start = current_end + timedelta(seconds=1)
    return costs


def generate_notes() -> list[dict]:
    templates = [
        {
            "Title": "Daily smoking habits",
            "Content": (
                "Today was a fairly regular day. I noticed that I tend to smoke "
                "more during the afternoon, especially when I am working or feeling stressed."
            ),
            "Mood": 3,
        },
        {
            "Title": "A better day",
            "Content": (
                "I smoked less than usual today. The longest break was in the afternoon, "
                "and it was easier than expected to avoid smoking."
            ),
            "Mood": 4,
        },
        {
            "Title": "Weekend",
            "Content": (
                "Smoking was more frequent today because I spent more time with friends. "
                "Most cigarettes were smoked in the evening."
            ),
            "Mood": 5,
        },
        {
            "Title": "Trying to reduce",
            "Content": (
                "I am trying to gradually reduce my daily consumption. "
                "The morning was easy, but I had more cravings later in the day."
            ),
            "Mood": 2,
        },
        {
            "Title": "Good progress",
            "Content": (
                "Today went well. I managed to keep longer intervals between cigarettes "
                "and felt that I had more control over my smoking."
            ),
            "Mood": 4,
        },
    ]

    notes = []

    now = datetime.now().replace(microsecond=0)
    for template in sample(templates, len(templates)):
        created_at = now - timedelta(
            days=randint(0, DAYS - 1),
            hours=randint(0, 23),
            minutes=randint(0, 59),
            seconds=randint(0, 59),
        )

        updated_at = created_at + timedelta(minutes=randint(1, 60))

        notes.append({
            **template,
            "CreatedAt": created_at.strftime(DATE_FORMAT),
            "UpdatedAt": updated_at.strftime(DATE_FORMAT),
        })
    return sorted(notes, key=lambda item: item["CreatedAt"])


def generate_settings() -> dict:
    currencies = ["€", "$", "£"]
    custom_currencies = ["CHF", "kr", "Kč", "zł", "¥"]

    return {
        "Theme": 1,
        "Language": 1,
        "Frequency": randint(0, 2),
        "Currency": choice(currencies),
        "CustomCurrency": choice(custom_currencies) if randint(0, 2) == 0 else "",
    }


def generate_notifications() -> dict:
    values = ["TRUE", "FALSE"]

    return {
        "System": choice(values),
        "Achievements": choice(values),
        "Progress": choice(values),
    }


def export_to_excel(data: dict[str, list[dict]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, rows in data.items():
        worksheet = workbook.create_sheet(sheet_name)

        if not rows:
            continue

        headers = list(rows[0].keys())
        worksheet.append(headers)

        for row in rows:
            worksheet.append([row.get(header) for header in headers])

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for column_index, header in enumerate(headers, start=1):
            column_letter = get_column_letter(column_index)

            max_length = max(
                len(str(worksheet.cell(row=row, column=column_index).value or ""))
                for row in range(1, worksheet.max_row + 1)
            )
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    workbook.save(OUTPUT_FILE)


def main() -> None:
    print("\n".join((
        "=" * 60,
        "DUMMY DATA GENERATOR",
        "=" * 60,
        "",
        f"Project:  {PROJECT_ROOT}",
        f"Output:   {OUTPUT_FILE}"
    )))

    history = generate_history()
    achievements = generate_achievements(history=history)
    costs = generate_costs()
    notes = generate_notes()
    settings = generate_settings()
    notifications = generate_notifications()

    export_to_excel({
        "History": history,
        "Achievements": achievements,
        "Costs": costs,
        "Notes": notes,
        "Settings": [settings],
        "Notifications": [notifications],
    })

    statistics = Statistics(history=len(history), achievements=len(achievements), costs=len(costs), notes=len(notes))
    statistics.summary()


if __name__ == "__main__":
    main()
